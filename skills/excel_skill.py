"""Excel skill pack — deep automation of Microsoft Excel."""

from tools.base_tool import BaseTool


class ExcelReadCellsTool(BaseTool):
    name = "excel_read_cells"
    description = "Read specific cells or ranges from an Excel file. Returns values as text."
    parameters = {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string", "default": ""},
        "range": {"type": "string", "description": "e.g. A1:D10 or A1"},
    }, "required": ["path"]}

    def run(self, path: str, sheet: str = "", range: str = "") -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            if range:
                rows = []
                for row in ws[range]:
                    if hasattr(row, '__iter__'):
                        rows.append(" | ".join(str(c.value or "") for c in row))
                    else:
                        return str(row.value)
                return "\n".join(rows)
            # Return first 20 rows
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20: break
                rows.append(" | ".join(str(v or "") for v in row))
            return f"Sheet: {ws.title}\n" + "\n".join(rows)
        except ImportError:
            return "Requires: pip install openpyxl"
        except Exception as e:
            return f"Error: {e}"


class ExcelWriteCellTool(BaseTool):
    name = "excel_write_cell"
    description = "Write a value to a specific cell in an Excel file."
    parameters = {"type": "object", "properties": {
        "path": {"type": "string"}, "cell": {"type": "string", "description": "e.g. B3"},
        "value": {}, "sheet": {"type": "string", "default": ""},
    }, "required": ["path", "cell", "value"]}

    def run(self, path: str, cell: str, value, sheet: str = "") -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            ws[cell] = value
            wb.save(path)
            return f"Written {value} to {cell}"
        except Exception as e:
            return f"Error: {e}"


class ExcelRunFormulaTool(BaseTool):
    name = "excel_formula"
    description = "Write an Excel formula to a cell (e.g. =SUM(A1:A10))."
    parameters = {"type": "object", "properties": {
        "path": {"type": "string"}, "cell": {"type": "string"},
        "formula": {"type": "string"}, "sheet": {"type": "string", "default": ""},
    }, "required": ["path", "cell", "formula"]}

    def run(self, path: str, cell: str, formula: str, sheet: str = "") -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            ws[cell] = formula
            wb.save(path)
            return f"Formula {formula} written to {cell}"
        except Exception as e:
            return f"Error: {e}"
